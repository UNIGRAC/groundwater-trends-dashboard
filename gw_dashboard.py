#!/usr/bin/env python
# coding: utf-8

# In[2]:


# Trial 3 with 2 pages (ALL VISUAL EDITS APPLIED) + Edit 1 (percentile fix) + Edit 2 (Page 2 red metadata items)
# ============================================================
# Dashboard + Metadata + 2-page HTML (tabs) + Indicator selector
# Edits included:
#   1) Bigger context box (Type/Name of spatial unit + wells counts)
#   2) Clearer indicator names
#   3) Clear legends (category + range + color) for pie + map (dynamic)
#   4) Improve hydrograph look (NO IQR band, NO smoothed mean; lighter clutter)
#   Edit 1) Percentile ranking FIXED: per-well, 2023 vs its own historical distribution (reference years before 2023)
#   Edit 2) Page 2: items shown in red are auto-adjusted (measurement type + units + Sen slope units/year in red, spelled out)
# ============================================================

import os, calendar, math, warnings
from pathlib import Path
from datetime import datetime

import numpy as np
import sys
import pandas as pd
from tqdm import tqdm

import pymannkendall as mk

import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.io as pio

warnings.simplefilter(action="ignore", category=FutureWarning)
warnings.filterwarnings("ignore")


# ============================================================
# Time-aware trendline builder (better for quarterly/gaps)
# ============================================================
def build_timeaware_trendline(dates: pd.DatetimeIndex, y: np.ndarray, slope_m_per_year: float):
    """Build y_trend using REAL elapsed time (years), not point index."""
    if len(dates) < 2 or slope_m_per_year is None or not np.isfinite(slope_m_per_year):
        return None
    dates = pd.to_datetime(dates)
    x_years = (dates - dates[0]).days / 365.25
    y = np.asarray(y, dtype=float)
    intercept = float(np.nanmean(y) - slope_m_per_year * np.nanmean(x_years))
    return slope_m_per_year * x_years + intercept


# ============================================================
# 2023 percentile ranking categories (per-well: 2023 vs its own history)
# ============================================================
def compute_2023_percentile_categories(
    accepted_df: pd.DataFrame,
    value_type: str,
    period_start: str = "2004-01-01",
    period_end: str = "2023-12-31",
):
    """
    Classify wells by the percentile of the TARGET YEAR (default = year in period_end)
    compared to the SAME WELL annual means in the reference period (years before target).

    - If depth: convert to "water table direction" (multiply by -1) so higher = higher water.
    - If elevation: use as is.
    """
    # convert to water-table direction for meaning consistency
    data = -accepted_df.copy() if str(value_type).strip().lower() == "depth" else accepted_df.copy()

    annual = data.resample("Y").mean()

    try:
        target_year = int(str(period_end)[:4])
    except Exception:
        target_year = 2023

    target_stamp = pd.Timestamp(f"{target_year}-12-31")

    # 2023 annual means per well
    try:
        vals_target = annual.loc[target_stamp]
    except KeyError:
        return pd.Series("No data in 2023", index=accepted_df.columns)

    # reference years: from period_start year up to (target_year - 1)
    try:
        ref_start_year = int(str(period_start)[:4])
    except Exception:
        ref_start_year = target_year - 19

    ref = annual[(annual.index.year >= ref_start_year) & (annual.index.year <= (target_year - 1))]

    # compute ECDF percentile per well using ONLY historical values
    def ecdf_percentile(hist: np.ndarray, v: float):
        hist = np.asarray(hist, dtype=float)
        hist = hist[np.isfinite(hist)]
        if hist.size == 0 or not np.isfinite(v):
            return np.nan
        # mid-rank ECDF for ties
        n = hist.size
        lt = np.sum(hist < v)
        eq = np.sum(hist == v)
        return (lt + 0.5 * eq) / n

    pct = pd.Series(index=accepted_df.columns, dtype=float)
    for well in accepted_df.columns:
        v = vals_target.get(well, np.nan)
        hist = ref[well].dropna().to_numpy() if well in ref.columns else np.array([])
        pct[well] = ecdf_percentile(hist, v)

    def classify(p):
        if pd.isna(p):
            return "No data in 2023"
        if p <= 0.10:
            return "Much below normal"
        if p < 0.25:
            return "Below normal"
        if p <= 0.75:
            return "Normal"
        if p < 0.90:
            return "Above normal"
        return "Much above normal"

    return pct.apply(classify)


# ============================================================
# Metadata CSV (field,value)
# ============================================================
def read_metadata_csv(path: Path) -> dict:
    """
    Reads a metadata CSV in the format:
        field,value
        Type of unit,Province
        Country,France
        Name of unit,Aquitaine
        Number of wells,120
        Accepted wells,85
        Type of measurements,depth
        Units,m b.g.l.
    """
    if path is None or (not Path(path).exists()):
        return {}
    df = pd.read_csv(path, encoding="utf-8")
    if df.shape[1] < 2:
        return {}
    df = df.iloc[:, :2]
    df.columns = ["field", "value"]
    out = {}
    for _, r in df.iterrows():
        k = str(r["field"]).strip()
        v = "" if pd.isna(r["value"]) else str(r["value"]).strip()
        if k:
            out[k] = v
    return out


def metadata_get(meta: dict, key: str, default=""):
    v = meta.get(key, default)
    return (v if v is not None else default)


# ============================================================
# Helper: detect value column + type (depth vs elevation)
# ============================================================
def detect_value_column_and_type(df: pd.DataFrame):
    """
    Detect which column holds groundwater values and infer type:
      depth (m b.g.l.) or elevation (m a.m.s.l.)
    """
    col_type_map = [
        ("depth", "depth"),
        ("water_level", "elevation"),
        ("waterlevel", "elevation"),
        ("gwl", "elevation"),
        ("elev", "elevation"),
        ("amsl", "elevation"),
        ("level", "depth"),  # fallback
    ]
    for key, vtype in col_type_map:
        for col in df.columns:
            if key in col.lower():
                return col, vtype
    raise ValueError(
        "Could not detect groundwater value column.\n"
        "Expected a column name containing: 'depth', 'water_level', 'waterlevel', 'gwl', 'elev', 'amsl', or 'level'."
    )


# ============================================================
# Data completeness filter
# ============================================================
def New_method_filter(
    df,
    percent=80,
    start="2004-01-01",
    end="2023-12-31",
    output_excel_path="Rejected_Wells.xlsx",
):
    start = datetime.strptime(start, "%Y-%m-%d")
    end = datetime.strptime(end, "%Y-%m-%d")

    df = df[(df.index >= start) & (df.index <= end)]
    total_wells = len(df.columns)

    n_months_period = (end.year - start.year) * 12 + (end.month - start.month) + 1
    total_years_period = n_months_period / 12.0
    total_years_required = total_years_period * (percent / 100.0)

    print(f"The chosen period covers {total_years_period} years between {start:%Y-%m-%d} and {end:%Y-%m-%d}")
    print(
        f"With the {percent}% condition this period will accept "
        f"{np.ceil(total_years_required)} ({round(total_years_required, 1)}) of {total_years_period}"
    )

    accepted_series_list = []
    rest_series_list = []
    rejected_wells_info = []

    for col_idx, well_name in tqdm(enumerate(df.columns), total=total_wells, desc="Completeness filter"):

        time_series = df.iloc[:, col_idx].reset_index()
        time_series_mc = time_series.groupby(time_series["Date"].dt.month_name(), sort=False).count()
        months_sorted = (
            time_series_mc.drop(columns="Date")
            .sort_values(by=well_name, ascending=False)
            .index
        )

        month_numbers = [list(calendar.month_name).index(m) for m in months_sorted]

        month_complete = []
        rest = [df.iloc[:, col_idx]]

        for i, month_num in enumerate(month_numbers):

            comp_dates = pd.DataFrame(
                pd.date_range(
                    start=start + pd.DateOffset(months=-1),
                    end=end + pd.DateOffset(months=+1),
                    freq="ME",
                ),
                columns=["Date"],
            )

            merged = comp_dates.merge(rest[i].reset_index(), how="left", on="Date")
            rest[i] = pd.Series(
                merged.iloc[:, 1].values,
                index=merged.iloc[:, 0].values,
                name=merged.iloc[:, 1].name,
            )
            rest[i].index.name = "Date"

            remaining = rest[i][rest[i].index.month != month_num]

            sel_period = rest[i][(rest[i].index >= start) & (rest[i].index <= end)]
            month_sel = sel_period[sel_period.index.month == month_num]

            month_sel_nan = month_sel[month_sel.isnull()]
            month_sel_nonan = month_sel.dropna()

            if len(month_sel_nonan) == total_years_period:
                month_complete.append(month_sel_nonan)
                rest.append(remaining)

            elif len(month_sel_nonan) < total_years_period:
                valores_reemplazo = [month_sel_nonan]

                for k, _ in enumerate(month_sel_nan):
                    idx_nan = month_sel_nan.index[k]

                    prev_idx = idx_nan + pd.DateOffset(months=-1) + pd.offsets.MonthEnd(n=0)
                    next_idx = idx_nan + pd.DateOffset(months=+1) + pd.offsets.MonthEnd(n=0)

                    prev_val = remaining.get(prev_idx, np.nan)
                    next_val = remaining.get(next_idx, np.nan)

                    if not np.isnan(prev_val) and not np.isnan(next_val):
                        month_sel_nan[k] = np.mean([prev_val, next_val])

                    elif not np.isnan(prev_val):
                        value = remaining.loc[[prev_idx]]
                        valores_reemplazo.append(value)
                        remaining = remaining.drop(prev_idx)

                    elif not np.isnan(next_val):
                        value = remaining.loc[[next_idx]]
                        valores_reemplazo.append(value)
                        remaining = remaining.drop(next_idx)

                paso1 = pd.concat(valores_reemplazo).sort_index()

                if len(pd.concat([paso1, month_sel_nan.dropna()]).sort_index()) >= total_years_required:
                    month_complete.append(pd.concat([paso1, month_sel_nan.dropna()]).sort_index())
                    rest.append(remaining)
                else:
                    rest.append(rest[i])

        comp_dates_full = pd.DataFrame(pd.date_range(start=start, end=end, freq="M"), columns=["Date"])

        if len(month_complete) > 0:
            time_series_nonan = pd.concat(month_complete).sort_index().reset_index()
            final_time_series = comp_dates_full.merge(time_series_nonan, how="left", on="Date").set_index("Date")
            rest_time_series = comp_dates_full.merge(rest[-1], how="left", on="Date").set_index("Date")
            accepted_series_list.append(final_time_series)
            rest_series_list.append(rest_time_series)
        else:
            rest_time_series = comp_dates_full.merge(rest[-1], how="left", on="Date").set_index("Date")
            rest_series_list.append(rest_time_series)
            rejected_wells_info.append({"Well": well_name, "Reason": "Insufficient Data"})

    if len(accepted_series_list) > 0:
        print(f"{len(accepted_series_list)} well(s) accepted out of {total_wells}")
        Final_accepted = pd.concat(accepted_series_list, axis=1)
        Not_accepted = pd.concat(rest_series_list, axis=1)
    else:
        print(f"0 wells accepted out of {total_wells}")
        Final_accepted = pd.DataFrame()
        Not_accepted = pd.concat(rest_series_list, axis=1)

    pd.DataFrame(rejected_wells_info).to_excel(output_excel_path, index=False)
    return Final_accepted, Not_accepted


# ============================================================
# Helper: interpret MK result as WATER TABLE trend
# ============================================================
def interpret_mk_result(result, value_type: str):
    """
    mk trend is for the raw series:
      - if value_type == depth, increasing depth = declining water table -> invert
      - if elevation, use as is
    """
    trend_raw = result.trend
    slope_raw = result.slope

    if value_type == "depth":
        if trend_raw == "increasing":
            trend = "decreasing"
        elif trend_raw == "decreasing":
            trend = "increasing"
        else:
            trend = trend_raw
        slope = -slope_raw
    else:
        trend = trend_raw
        slope = slope_raw

    return trend, slope


# ============================================================
# MK tests (20-year)
# ============================================================
def MK_tests(df, value_type="depth", alpha=0.1):
    print("Mann-Kendall Test to execute: mk.hamed_rao_modification_test")
    print(f"Value type interpreted as: {value_type}")

    results_list = []
    n_months = len(df.index.unique())

    for i, well_name in enumerate(df.columns):
        series = df.iloc[:, i]
        result = mk.hamed_rao_modification_test(series, alpha=alpha)
        trend, slope_per_month = interpret_mk_result(result, value_type=value_type)

        slope_per_year = slope_per_month * 12.0
        slope_total_period = slope_per_month * n_months

        results_list.append(
            {
                "Well": well_name,
                "Trend": trend,
                "p": result.p,
                "mk_slope_20yr_m_per_month": slope_per_month,
                "mk_slope_20yr_m_per_year": slope_per_year,
                "slope_total_period": slope_total_period,
            }
        )

    return pd.DataFrame(results_list)


# ============================================================
# MK slopes for fixed 10 / 5-year windows (monthly-based)
# ============================================================
def compute_mk_slopes_fixed_windows_monthly(df, value_type, alpha=0.1, percent=80):
    windows = {
        "mk_slope_10yr_m_per_year": (pd.Timestamp("2014-01-01"), pd.Timestamp("2023-12-31"), 10),
        "mk_slope_5yr_m_per_year": (pd.Timestamp("2019-01-01"), pd.Timestamp("2023-12-31"), 5),
    }

    rows = []
    for well in df.columns:
        row = {"Well": well}
        series = df[well]

        for col_name, (p_start, p_end, _) in windows.items():
            sub = series[(series.index >= p_start) & (series.index <= p_end)]

            total_months_window = len(pd.date_range(start=p_start, end=p_end, freq="ME"))
            n_months_with_data = sub.dropna().shape[0]
            required_months = math.ceil(total_months_window * (percent / 100.0))

            if n_months_with_data < required_months or n_months_with_data < 3:
                row[col_name] = np.nan
                continue

            try:
                result = mk.hamed_rao_modification_test(sub.dropna(), alpha=alpha)
            except ZeroDivisionError:
                row[col_name] = np.nan
                continue

            _, slope_per_month = interpret_mk_result(result, value_type=value_type)
            row[col_name] = slope_per_month * 12.0

        rows.append(row)

    return pd.DataFrame(rows)


# ============================================================
# DASHBOARD: 2 pages (tabs)
# ============================================================
def build_dashboard(
    df_mk: pd.DataFrame,
    accepted_df: pd.DataFrame,
    value_type: str,
    units_str: str,
    output_html: Path,
    region_name: str,
    metadata: dict,
    period_start="2004-01-01",
    period_end="2023-12-31",
    completeness_percent=80,
    alpha=0.1,
):
    # ---------------------------
    # Shared color palettes
    # ---------------------------
    cri_colors = {
        "Strong declining": "#FF5900",
        "Moderate declining": "#FFA600",
        "Weak slope": "#747474",
        "Moderate rising": "#87D6E6",
        "Strong rising": "#007BFF",
        "No data": "#B0B0B0",
    }

    mk_trend_colors = {
        "Significant Decrease": "#FF5900",
        "No trend": "#747474",
        "Significant Increase": "#007BFF",
        "No data": "#B0B0B0",
    }

    rank_colors = {
        "Much below normal": "#FF5900",
        "Below normal": "#FFA600",
        "Normal": "#747474",
        "Above normal": "#87D6E6",
        "Much above normal": "#007BFF",
        "No data in 2023": "#B0B0B0",
    }

    # ---------------------------
    # Counts (allow override from metadata, optional)
    # ---------------------------
    n_accepted = accepted_df.shape[1]
    n_with_coords = df_mk.dropna(subset=["X", "Y"]).shape[0] if {"X", "Y"}.issubset(df_mk.columns) else 0

    def _safe_int(s):
        try:
            return int(float(str(s).strip()))
        except Exception:
            return None

    n_input_override = _safe_int(metadata_get(metadata, "Number of wells", "").strip() or None)
    n_acc_override = _safe_int(metadata_get(metadata, "Accepted wells", "").strip() or None)

    n_input_computed = df_mk["Well"].nunique() if "Well" in df_mk else n_accepted
    n_input = n_input_override if (n_input_override is not None) else n_input_computed
    n_accepted_show = n_acc_override if (n_acc_override is not None) else n_accepted

    n_mk20 = df_mk["mk_slope_20yr_m_per_year"].notna().sum() if "mk_slope_20yr_m_per_year" in df_mk else 0
    n_cri = df_mk["CRI"].notna().sum() if "CRI" in df_mk else 0

    # Edit 1: FIXED percentile series (per-well, 2023 vs own history)
    rank_series = compute_2023_percentile_categories(
        accepted_df=accepted_df,
        value_type=value_type,
        period_start=period_start,
        period_end=period_end,
    )
    n_rank2023 = (rank_series != "No data in 2023").sum()

    n_senslope20 = n_mk20

    # ---------------------------
    # Hydrograph prep (NO IQR band, NO smoothing)
    # ---------------------------
    data_for_hydro = -accepted_df.copy() if value_type == "depth" else accepted_df.copy()
    data_for_hydro = data_for_hydro.dropna(how="all", axis=1)

    norm_df = data_for_hydro.sub(data_for_hydro.mean(axis=0), axis=1)
    mean_norm = norm_df.mean(axis=1)
    mean_norm_df = mean_norm.to_frame(name="mean_norm").sort_index()
    mean_norm_df.index.name = "Date"

    periods = {
        20: (pd.Timestamp("2004-01-01"), pd.Timestamp("2023-12-31")),
        10: (pd.Timestamp("2014-01-01"), pd.Timestamp("2023-12-31")),
        5: (pd.Timestamp("2019-01-01"), pd.Timestamp("2023-12-31")),
    }

    mean_slopes_year = {
        20: df_mk["mk_slope_20yr_m_per_year"].mean(skipna=True),
        10: df_mk["mk_slope_10yr_m_per_year"].mean(skipna=True) if "mk_slope_10yr_m_per_year" in df_mk else np.nan,
        5: df_mk["mk_slope_5yr_m_per_year"].mean(skipna=True) if "mk_slope_5yr_m_per_year" in df_mk else np.nan,
    }
    counts_year = {
        20: df_mk["mk_slope_20yr_m_per_year"].notna().sum() if "mk_slope_20yr_m_per_year" in df_mk else 0,
        10: df_mk["mk_slope_10yr_m_per_year"].notna().sum() if "mk_slope_10yr_m_per_year" in df_mk else 0,
        5: df_mk["mk_slope_5yr_m_per_year"].notna().sum() if "mk_slope_5yr_m_per_year" in df_mk else 0,
    }

    # ---------------------------
    # MK significance categories (pie + map)
    # ---------------------------
    def normalize_trend(t):
        if pd.isna(t):
            return "No data"
        t = str(t).strip().lower()
        if "inc" in t:
            return "Significant Increase"
        if "dec" in t:
            return "Significant Decrease"
        if "no" in t:
            return "No trend"
        return "No data"

    df_mk = df_mk.copy()
    df_mk["MK_trend_category"] = df_mk["Trend"].apply(normalize_trend)

    mk_trend_order = ["Significant Decrease", "No trend", "Significant Increase", "No data"]
    mk_trend_counts = df_mk["MK_trend_category"].value_counts().reindex(mk_trend_order, fill_value=0)

    # ---------------------------
    # CRI categories (pie + map)
    # ---------------------------
    def classify_cri(cri):
        if pd.isna(cri):
            return "Not defined (IQR=0)"
        if cri > 1:
            return "Strong rising"
        if cri > 0.5:
            return "Moderate rising"
        if cri > -0.5:
            return "Weak slope"
        if cri > -1:
            return "Moderate declining"
        return "Strong declining"

    df_mk["CRI_category"] = df_mk["CRI"].apply(classify_cri) if "CRI" in df_mk else "No data"
    cri_order = ["Strong declining", "Moderate declining", "Weak slope", "Moderate rising", "Strong rising", "No data"]
    cri_counts = df_mk["CRI_category"].value_counts().reindex(cri_order, fill_value=0)

    # ---------------------------
    # 2023 ranking categories (pie + map)
    # ---------------------------
    df_mk["Rank2023_category"] = df_mk["Well"].map(rank_series)
    rank_order = ["Much below normal", "Below normal", "Normal", "Above normal", "Much above normal", "No data in 2023"]
    rank_counts = df_mk["Rank2023_category"].value_counts().reindex(rank_order, fill_value=0)

    # ---------------------------
    # Sen slope (20y m/yr) classes (pie + map)
    # ---------------------------
    def classify_sen_slope_myr(s):
        if pd.isna(s):
            return "No data"
        if s < -0.25:
            return "Strong declining"
        if s < -0.1:
            return "Moderate declining"
        if s <= 0.1:
            return "Weak slope"
        if s <= 0.25:
            return "Moderate rising"
        return "Strong rising"

    df_mk["SenSlope20_category"] = df_mk["mk_slope_20yr_m_per_year"].apply(classify_sen_slope_myr)
    senslope_counts = df_mk["SenSlope20_category"].value_counts().reindex(cri_order, fill_value=0)

    # ---------------------------
    # Map data
    # ---------------------------
    if not {"X", "Y"}.issubset(df_mk.columns):
        raise ValueError("MK_results must contain 'X' and 'Y' columns with lon/lat (EPSG:4326).")

    df_map = df_mk.dropna(subset=["X", "Y"]).copy()

    df_map["hover_text"] = (
        "Well: " + df_map["Well"].astype(str)
        + "<br>MK (20y): " + df_map["MK_trend_category"].astype(str)
        + "<br>CRI: " + (df_map["CRI"].round(3).astype(str) if "CRI" in df_map else "NA")
        + "<br>CRI category: " + df_map["CRI_category"].astype(str)
        + "<br>2023 status: " + df_map["Rank2023_category"].astype(str)
        + "<br>Sen slope 20y (m/yr): " + df_map["mk_slope_20yr_m_per_year"].round(3).astype(str)
        + "<br>Sen slope class: " + df_map["SenSlope20_category"].astype(str)
    )

    # ---------------------------
    # Figure layout: pie + map + hydrograph
    # ---------------------------
    fig = make_subplots(
        rows=3,
        cols=1,
        specs=[
            [{"type": "domain"}],
            [{"type": "geo"}],
            [{"type": "xy"}],
        ],
        row_heights=[0.28, 0.44, 0.28],
        vertical_spacing=0.1,
    )

    # ---------------------------
    # Context box (BIGGER + includes Type/Name/Counts)
    # ---------------------------
    unit_type = metadata_get(metadata, "Type of unit", "").strip()
    country = metadata_get(metadata, "Country", "").strip()
    unit_name = metadata_get(metadata, "Name of unit", region_name).strip()
    meas_type_forced = metadata_get(metadata, "Type of measurements", value_type).strip()
    units_forced = metadata_get(metadata, "Units", units_str).strip()
    u = units_forced.strip().lower()

    if u in ["m b.g.l.", "mbgl", "m bgl", "m below ground level", "meters below ground level", "metres below ground level"]:
        units_display = "meters below ground level"
    elif u in ["m a.m.s.l.", "mamsl", "m amsl", "m above mean sea level", "meters above mean sea level", "metres above mean sea level"]:
        units_display = "meters above mean sea level"
    else:
        units_display = units_forced

    context_lines = [
        "<span style='font-size:14px;'><b>CONTEXT</b></span>",
        f"<b>Type of spatial unit:</b> {unit_type}" if unit_type else "",
        f"<b>Name of spatial unit:</b> {unit_name}",
        f"<b>Country:</b> {country}" if country else "",
        f"<b>Measurement type:</b> {meas_type_forced}",
        f"<b>Measurement unit:</b> {units_display}",
        f"<b>Number of wells:</b> {n_input} input → <b>{n_accepted_show}</b> accepted",
        f"<b>Wells with coordinates:</b> {n_with_coords}",
        f"<b>Period:</b> {period_start} to {period_end} ",
    ]
    context_text = "<br>".join([x for x in context_lines if x])

    # ---------------------------
    # Clearer indicator names
    # ---------------------------
    IND_MK = "Mann–Kendall significance (20y)"
    IND_CRI = "Cumulative Relative Impact (CRI)"
    IND_RANK = "Groundwater status in 2023 (percentile)"
    IND_SEN = "Sen slope classes (20y, m/yr)"
    indicator_titles = [IND_MK, IND_CRI, IND_RANK, IND_SEN]

    indicator_nline = {
        IND_MK: f"Wells used for Mann–Kendall significance (20y): <b>{n_mk20}</b>",
        IND_CRI: f"Wells used for CRI calculation: <b>{n_cri}</b>",
        IND_RANK: f"Wells used for percentile calculation (2023): <b>{n_rank2023}</b>",
        IND_SEN: f"Wells used for Sen slope classes (20y): <b>{n_senslope20}</b>",
    }

    # ---------------------------
    # Legends (category + range + color) shown as annotation box (dynamic)
    # ---------------------------
    def _legend_rows(items):
        rows = []
        for lab, rng, col in items:
            if str(rng).strip():
                rows.append(f"<span style='color:{col}; font-size:16px;'><b>■</b></span> <b>{lab}</b>: {rng}")
            else:
                rows.append(f"<span style='color:{col}; font-size:16px;'><b>■</b></span> <b>{lab}</b>")
        return "<br>".join(rows)

    legend_defs = {
        IND_MK: _legend_rows([
            ("Significant Decrease", "", mk_trend_colors["Significant Decrease"]),
            ("No trend", "", mk_trend_colors["No trend"]),
            ("Significant Increase", "", mk_trend_colors["Significant Increase"]),
        ]),
        IND_CRI: _legend_rows([
            ("Strong declining", "", cri_colors["Strong declining"]),
            ("Moderate declining", "", cri_colors["Moderate declining"]),
            ("Weak slope", "", cri_colors["Weak slope"]),
            ("Moderate rising", "", cri_colors["Moderate rising"]),
            ("Strong rising", "", cri_colors["Strong rising"]),
            ("Not defined (IQR=0)", "", cri_colors["No data"]),
        ]),
        IND_RANK: _legend_rows([
            ("Much below normal", "", rank_colors["Much below normal"]),
            ("Below normal", "", rank_colors["Below normal"]),
            ("Normal", "", rank_colors["Normal"]),
            ("Above normal", "", rank_colors["Above normal"]),
            ("Much above normal", "", rank_colors["Much above normal"]),
            ("No data in 2023", "", rank_colors["No data in 2023"]),
        ]),
        IND_SEN: _legend_rows([
            ("Strong declining", "", cri_colors["Strong declining"]),
            ("Moderate declining", "", cri_colors["Moderate declining"]),
            ("Weak slope", "", cri_colors["Weak slope"]),
            ("Moderate rising", "", cri_colors["Moderate rising"]),
            ("Strong rising", "", cri_colors["Strong rising"]),
        ]),
    }

    def make_annotations(indicator_title: str):
        legend_html = "<span style='font-size:14px;'><b>LEGEND</b></span><br>" + legend_defs.get(indicator_title, "")
        return [
            dict(
                x=0.5, y=1.065, xref="paper", yref="paper",
                text=f"<b>Groundwater Trends Dashboard – {unit_name} ({n_accepted_show} wells)</b>",
                showarrow=False, font=dict(family="Arial Black", size=22), xanchor="center"
            ),
            dict(
                x=0.02, y=0.92, xref="paper", yref="paper",
                text=f"<b>Distribution of wells</b><br>{indicator_title}",
                showarrow=False, font=dict(family="Arial Black", size=13), xanchor="left", align="left"
            ),
            dict(
                x=0.02, y=0.60, xref="paper", yref="paper",
                text=f"<b>Spatial distribution</b><br>{indicator_title}",
                showarrow=False, font=dict(family="Arial Black", size=13), xanchor="left", align="left"
            ),
            dict(
                x=0.02, y=0.26, xref="paper", yref="paper",
                text="<b>Temporal context</b> — Mean normalized hydrograph & trendlines",
                showarrow=False, font=dict(family="Arial Black", size=13), xanchor="left", yanchor="bottom", align="left"
            ),
            dict(
                x=1.02, y=0.24, xref="paper", yref="paper",
                text="<i>Tip:</i> Zoom, pan, or double-click to reset the time range",
                showarrow=False, font=dict(family="Arial", size=11, color="#555"),
                xanchor="left", align="left"
            ),
            dict(
                x=0.995, y=0.985, xref="paper", yref="paper",
                text=context_text,
                showarrow=False, align="left",
                xanchor="right", yanchor="top",
                font=dict(family="Arial", size=13),
                bgcolor="rgba(255,255,255,0.96)",
                bordercolor="rgba(0,0,0,0.25)",
                borderwidth=2
            ),
            dict(
                x=1.03, y=0.5, xref="paper", yref="paper",
                text="<i>Note:</i> For more information, check <b><br>Page 2 — Methodology</b>",
                showarrow=False,
                align="left",
                xanchor="left",
                yanchor="top",
                font=dict(family="Arial", size=11, color="#555"),
            ),
            dict(
                x=0.02, y=0.885, xref="paper", yref="paper",
                text=f"<span style='color:#333;'>{indicator_nline.get(indicator_title,'')}</span>",
                showarrow=False, font=dict(family="Arial", size=11),
                xanchor="left", align="left"
            ),
            dict(
                x=1.03, y=0.62, xref="paper", yref="paper",
                text=legend_html,
                showarrow=False, align="left",
                xanchor="left", yanchor="top",
                font=dict(family="Arial", size=12),
                bgcolor="rgba(255,255,255,0.96)",
                bordercolor="rgba(0,0,0,0.25)",
                borderwidth=2
            ),
        ]

    # ---------------------------
    # Pie traces (4, 1 visible)
    # ---------------------------
    pies = [
        go.Pie(
            labels=mk_trend_counts.index, values=mk_trend_counts.values,
            textinfo="label+percent", textposition="inside",
            insidetextorientation="radial",
            marker=dict(colors=[mk_trend_colors[l] for l in mk_trend_counts.index], line=dict(width=1, color="white")),
            textfont=dict(color="white", size=12, family="Arial"),
            showlegend=False, visible=True
        ),
        go.Pie(
            labels=cri_counts.index, values=cri_counts.values,
            textinfo="label+percent", textposition="inside",
            insidetextorientation="radial",
            marker=dict(colors=[cri_colors[l] for l in cri_counts.index], line=dict(width=1, color="white")),
            textfont=dict(color="white", size=12, family="Arial"),
            showlegend=False, visible=False
        ),
        go.Pie(
            labels=rank_counts.index, values=rank_counts.values,
            textinfo="label+percent", textposition="inside",
            insidetextorientation="radial",
            marker=dict(colors=[rank_colors[l] for l in rank_counts.index], line=dict(width=1, color="white")),
            textfont=dict(color="white", size=12, family="Arial"),
            showlegend=False, visible=False
        ),
        go.Pie(
            labels=senslope_counts.index, values=senslope_counts.values,
            textinfo="label+percent", textposition="inside",
            insidetextorientation="radial",
            marker=dict(colors=[cri_colors[l] for l in senslope_counts.index], line=dict(width=1, color="white")),
            textfont=dict(color="white", size=12, family="Arial"),
            showlegend=False, visible=False
        ),
    ]
    for tr in pies:
        fig.add_trace(tr, row=1, col=1)

    # ---------------------------
    # Map traces (4, 1 visible)
    # ---------------------------
    maps = [
        go.Scattergeo(
            lon=df_map["X"], lat=df_map["Y"], text=df_map["hover_text"],
            mode="markers",
            marker=dict(size=8, color=df_map["MK_trend_category"].map(mk_trend_colors).fillna("#B0B0B0"),
                        line=dict(width=0.6, color="black")),
            showlegend=False, visible=True
        ),
        go.Scattergeo(
            lon=df_map["X"], lat=df_map["Y"], text=df_map["hover_text"],
            mode="markers",
            marker=dict(size=8, color=df_map["CRI_category"].map(cri_colors).fillna("#B0B0B0"),
                        line=dict(width=0.6, color="black")),
            showlegend=False, visible=False
        ),
        go.Scattergeo(
            lon=df_map["X"], lat=df_map["Y"], text=df_map["hover_text"],
            mode="markers",
            marker=dict(size=8, color=df_map["Rank2023_category"].map(rank_colors).fillna("#B0B0B0"),
                        line=dict(width=0.6, color="black")),
            showlegend=False, visible=False
        ),
        go.Scattergeo(
            lon=df_map["X"], lat=df_map["Y"], text=df_map["hover_text"],
            mode="markers",
            marker=dict(size=8, color=df_map["SenSlope20_category"].map(cri_colors).fillna("#B0B0B0"),
                        line=dict(width=0.6, color="black")),
            showlegend=False, visible=False
        ),
    ]
    for tr in maps:
        fig.add_trace(tr, row=2, col=1)

    fig.update_geos(
        fitbounds="locations",
        showcountries=True,
        showland=True,
        landcolor="rgb(240,240,240)",
        lakecolor="rgb(255,255,255)",
        showlakes=True,
        row=2, col=1
    )

    fig.layout.geo.domain.x = [0.2, 0.98]

    # ---------------------------
    # Hydrograph (clean look, NO IQR band, NO smoothing)
    # ---------------------------
    cols = list(norm_df.columns)
    max_spaghetti = 200
    if len(cols) > max_spaghetti:
        rng = np.random.default_rng(42)
        cols = list(rng.choice(cols, size=max_spaghetti, replace=False))

    for c in cols:
        fig.add_trace(
            go.Scatter(
                x=norm_df.index,
                y=norm_df[c],
                mode="lines",
                line=dict(width=1, color="rgba(120,120,120,0.10)"),
                hoverinfo="skip",
                showlegend=False,
            ),
            row=3, col=1
        )

    fig.add_trace(
        go.Scatter(
            x=mean_norm_df.index,
            y=mean_norm_df["mean_norm"],
            mode="lines",
            line=dict(width=3, color="rgba(0,0,0,0.65)"),
            name="Mean groundwater level (all wells, monthly)",
            showlegend=True,
        ),
        row=3, col=1
    )

    colors_periods = {20: "blue", 10: "green", 5: "red"}
    for yrs, (p_start, p_end) in periods.items():
        slope_yr = mean_slopes_year.get(yrs, np.nan)
        if not np.isfinite(slope_yr):
            continue

        sub = mean_norm_df[(mean_norm_df.index >= p_start) & (mean_norm_df.index <= p_end)]
        if len(sub) < 2:
            continue

        y_trend = build_timeaware_trendline(sub.index, sub["mean_norm"].to_numpy(), slope_yr)
        if y_trend is None:
            continue

        n_wells_period = counts_year.get(yrs, 0)

        fig.add_trace(
            go.Scatter(
                x=sub.index,
                y=y_trend,
                mode="lines",
                line=dict(width=3, dash="dash", color=colors_periods[yrs]),
                name=f"{yrs}-yr MK Sen slope ({slope_yr:.3f} {units_str}/yr, n={n_wells_period})",
                showlegend=True,
            ),
            row=3, col=1
        )

    all_vals = norm_df.to_numpy().ravel()
    all_vals = all_vals[~np.isnan(all_vals)]
    y_range = None
    if len(all_vals) > 0:
        low_p, high_p = np.nanpercentile(all_vals, [2, 98])
        margin = max((high_p - low_p) * 0.12, 0.6)
        y_range = [low_p - margin, high_p + margin]

    fig.update_xaxes(title_text="Year", row=3, col=1, title_font=dict(family="Arial Black", size=12))
    fig.update_yaxes(
        title_text="Normalized water table (m)",
        row=3, col=1,
        range=y_range,
        zeroline=True,
        zerolinewidth=2,
        zerolinecolor="black",
        title_font=dict(family="Arial Black", size=12),
    )

    # ---------------------------
    # Dropdown: toggles pie+map + updates annotations (legend changes)
    # ---------------------------
    n_traces = len(fig.data)

    def visible_mask(which: int):
        mask = [False] * n_traces
        mask[which] = True
        mask[4 + which] = True
        for i in range(8, n_traces):
            mask[i] = True
        return mask

    buttons = []
    for i, t in enumerate(indicator_titles):
        buttons.append(dict(
            label=t,
            method="update",
            args=[{"visible": visible_mask(i)}, {"annotations": make_annotations(t)}],
        ))

    fig.update_layout(
        updatemenus=[dict(
            type="dropdown",
            direction="down",
            x=0.02, y=1.02,
            xanchor="left", yanchor="bottom",
            buttons=buttons,
            bgcolor="rgba(255,255,255,0.95)",
            bordercolor="rgba(0,0,0,0.15)",
            borderwidth=1,
            font=dict(family="Arial", size=12),
        )],
        annotations=make_annotations(indicator_titles[0]),
        margin=dict(l=40, r=280, t=95, b=35),
        height=1400,
        font=dict(family="Arial", size=11),
        hovermode="x unified",
        plot_bgcolor="white",
        legend=dict(
            title="Hydrograph",
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="rgba(0,0,0,0.20)",
            borderwidth=1,
            x=1.02, xanchor="left",
            y=0.02, yanchor="bottom"
        )
    )

    # ---------------------------
    # PAGE 2: Methodology (Edit 2 applied: meta-input red + Sen slope units spelled out in red)
    # ---------------------------
    methodology_pdf_url = "https://un-igrac.org/wp-content/uploads/2024/12/CN_Reporting-methodology_GW-levels-2025_final-1.1.pdf"
    page2_html = f"""
    <div class="page2">
      <h2>Methodology &amp; Data Coverage</h2>

      <h3>1. Input data</h3>
      <p>
        Groundwater level time series <b>should be provided</b> as <b>daily and/or monthly observations</b>.
        The <b>code aggregates</b> all input data to <b>monthly means</b> prior to analysis.
      </p>
      <p>
        Measurements represent groundwater
        <span class="meta-input">{meas_type_forced}</span>,
        expressed in
        <span class="meta-input">{units_display}</span>.
      </p>
      <div class="meta-note">
        <b>Note:</b> Items shown in red are automatically adjusted based on the metadata file.
      </div>

      <h3>2. Data completeness and selection</h3>
      <ul>
        <li>Analysis period: <b>{period_start}–{period_end}</b></li>
        <li>Minimum data availability: <b>{completeness_percent}%</b> of the period</li>
        <li>Only wells meeting the data availability threshold are included in the indicators.</li>
      </ul>
      <p>
        <b>Download:</b>
        <a href="{methodology_pdf_url}" target="_blank" rel="noopener noreferrer">
          Groundwater levels reporting methodology (PDF)
        </a>
      </p>

      <h3>3. Mann–Kendall trend and significance</h3>
      <p>
        The Mann–Kendall (MK) test evaluates whether there is a <b>monotonic trend</b> (consistently increasing or decreasing)
        in the groundwater time series over the selected period.
      </p>
      <ul>
        <li><b>Significant Increase / Decrease:</b> the trend is statistically detectable at the chosen significance level.</li>
        <li><b>No trend (not significant):</b> the test does not find enough evidence of a monotonic trend (this does not prove that no change exists).</li>
      </ul>

      <p>
        To account for serial autocorrelation in hydrological time series, a <b>modified Mann–Kendall test</b>
        (variance correction) is applied following Hamed &amp; Rao (1998).
        <br>
        Reference: Hamed, K.H. &amp; Rao, A.R. (1998)
        <i>A modified Mann–Kendall trend test for autocorrelated data</i>,
        <b>Journal of Hydrology</b>, 204(1–4), 182–196.
        DOI:
        <a href="https://doi.org/10.1016/S0022-1694(97)00125-X" target="_blank">
            10.1016/S0022-1694(97)00125-X
        </a>.
       </p>

      <h3>4. Sen’s slope (trend magnitude)</h3>
      <p>
        Sen’s slope estimates the <b>magnitude</b> of the trend, expressed in
        <span class="meta-input">{units_display}/year</span>.
        A negative slope indicates declining water levels (after converting depth to “water table” direction).
      </p>

      <h3>5. Time windows</h3>
      <ul>
        <li><b>20-year:</b> 2004–2023</li>
        <li><b>10-year:</b> 2014–2023</li>
        <li><b>5-year:</b> 2019–2023</li>
      </ul>

      <h3>6. Cumulative Relative Impact (CRI)</h3>
      <p>
        <b>CRI</b> expresses how large the total change is relative to typical variability in the well record:
        <br>
        <b>CRI = Total groundwater level change over the analysis period / Interquartile Range (IQR)</b>
      </p>
      <ul>
        <li><b>Strong declining:</b> CRI &le; −1</li>
        <li><b>Moderate declining:</b> −1 &lt; CRI &le; −0.5</li>
        <li><b>Weak slope:</b> −0.5 &lt; CRI &le; 0.5</li>
        <li><b>Moderate rising:</b> 0.5 &lt; CRI &le; 1</li>
        <li><b>Strong rising:</b> CRI &gt; 1</li>
      </ul>

      <h3>7. Groundwater status in 2023 (percentile)</h3>
      <p>
        The 2023 groundwater status compares the <b>2023 annual mean</b> of each well against the <b>historical distribution of the same well</b>
        (reference years before 2023) within the selected period.
        Percentiles describe whether a well is relatively low or high in 2023 compared with its own historical values.
      </p>
      <ul>
        <li><b>Much below normal:</b> &le; 10th percentile</li>
        <li><b>Below normal:</b> 10th–25th percentile</li>
        <li><b>Normal:</b> 25th–75th percentile</li>
        <li><b>Above normal:</b> 75th–90th percentile</li>
        <li><b>Much above normal:</b> &ge; 90th percentile</li>
      </ul>
      <hr style="margin-top:30px;">

        <p style="font-size:11px; color:#666; line-height:1.4;">
        © 2026 International Groundwater Resources Assessment Centre (IGRAC). All rights reserved.<br>
        This application is intended for informational and analytical purposes only.<br>
        IGRAC does not guarantee the completeness or accuracy of the results.<br>
        Underlying groundwater data remain the property of the original data providers.
        </p>

    </div>
    """

    # ---------------------------
    # Build 2-tab HTML
    # ---------------------------
    fig_html = pio.to_html(fig, include_plotlyjs="cdn", full_html=False, config={"displaylogo": False})

    combined_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="utf-8"/>
      <title>Groundwater Trends Dashboard – {unit_name}</title>
      <style>
        body {{ font-family: Arial, sans-serif; margin: 0; background: #ffffff; }}
        .topbar {{ display:flex; gap:8px; padding:10px 14px; border-bottom:1px solid rgba(0,0,0,0.12); align-items:center; }}
        .tabbtn {{ padding:8px 12px; border:1px solid rgba(0,0,0,0.15); border-radius:10px; background:rgba(255,255,255,0.95); cursor:pointer; font-weight:700; }}
        .tabbtn.active {{ background:rgba(0,123,255,0.10); border-color:rgba(0,123,255,0.35); }}
        .tabcontent {{ display:none; padding:10px 12px; }}
        .tabcontent.active {{ display:block; }}
        .page2 {{ max-width:980px; margin:0 auto; padding:16px 18px 40px 18px; line-height:1.5; }}
        .page2 h2 {{ margin-top:8px; font-size:28px; font-weight:800; }}
        .page2 h3 {{ margin-top:18px; font-size:18px; font-weight:800; }}
        .meta-input {{ color:#c00000; font-weight:800; }}
          .igrac-logo {{
            position: fixed;
            top: 12px;
            right: 18px;
            height: 44px;
            padding: 6px 10px;
            background: rgba(255,255,255,0.95);
            border-radius: 8px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.15);
            z-index: 9999;
            }}

      </style>
    </head>
    <body>
    <img
        src="../input/IGRAC_logo_FC.png"
        class="igrac-logo"
        alt="IGRAC logo"
    />

      <div class="topbar">
        <button class="tabbtn active" onclick="showTab('tab1', this)">Page 1 — Overview</button>
        <button class="tabbtn" onclick="showTab('tab2', this)">Page 2 — Methodology</button>
      </div>

      <div id="tab1" class="tabcontent active">
        {fig_html}
      </div>

      <div id="tab2" class="tabcontent">
        {page2_html}
      </div>

      <script>
        function showTab(tabId, btn) {{
          document.querySelectorAll('.tabcontent').forEach(el => el.classList.remove('active'));
          document.getElementById(tabId).classList.add('active');
          document.querySelectorAll('.tabbtn').forEach(b => b.classList.remove('active'));
          btn.classList.add('active');
        }}
      </script>
    </body>
    </html>
    """

    output_html = Path(output_html)
    output_html.write_text(combined_html, encoding="utf-8")
    print(f"HTML dashboard (2 pages) saved to: {output_html}")


# ============================================================
# MAIN PIPELINE – FROM RAW CSV TO HTML + MK_results.xlsx
# ============================================================
if __name__ == "__main__":

        # --------------------------------------------------------
    # USER SETTINGS (DO NOT EDIT PATHS)
    # --------------------------------------------------------

    # Base directory = folder where the script / exe is located
    if getattr(sys, 'frozen', False):
        # Running as compiled exe
        base_path = Path(sys.executable).parent
    else:
        # Running as .py script
        base_path = Path(__file__).resolve().parent

    input_dir = base_path / "input"
    output_dir = base_path / "output"

    input_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)

    # --------------------------------------------------------
    # INPUT FILES (must be placed in /input)
    # --------------------------------------------------------
    input_timeseries_file = input_dir / "Monitoring_data.csv"
    coord_file = input_dir / "Sites_coordinates.csv"

    metadata_file_1 = input_dir / f"{base_path.name}_metadata.csv"
    metadata_file_2 = input_dir / "metadata.csv"
    metadata_path = metadata_file_1 if metadata_file_1.exists() else metadata_file_2

    # --------------------------------------------------------
    # OUTPUT FILES (written to /output)
    # --------------------------------------------------------
    accepted_csv_path = output_dir / "Accepted_data.csv"
    rejected_excel_path = output_dir / "Rejected_Wells.xlsx"
    mk_results_path = output_dir / "MK_results.xlsx"
    hydrograph_png = output_dir / "Hydrograph_mean_normalized.png"
    output_html = output_dir / "Dashboard_2pages_Selector_MK_CRI_Rank_SenSlope.html"

    completeness_percent = 80
    period_start = "2004-01-01"
    period_end = "2023-12-31"
    alpha = 0.1

    # --------------------------------------------------------
    # B) READ METADATA (FORCED SEMANTICS)
    # --------------------------------------------------------
    metadata = read_metadata_csv(metadata_path)
    forced_measurement_type = metadata_get(metadata, "Type of measurements", "").strip().lower()
    forced_units = metadata_get(metadata, "Units", "m").strip()

    # --------------------------------------------------------
    # C) READ RAW TIME SERIES
    # --------------------------------------------------------
    raw_df = pd.read_csv(
        input_timeseries_file,
        encoding="utf-8",
        parse_dates=["Date"],
        dayfirst=True,
    )

    if "site" not in raw_df.columns:
        raise ValueError("Input time series CSV must contain a 'site' column (well identifier).")
    
        # --------------------------------------------------------
    # TOTAL INPUT WELLS (accepted + rejected)
    # --------------------------------------------------------
    p_start = pd.to_datetime(period_start)
    p_end   = pd.to_datetime(period_end)

    raw_df_period = raw_df[
        (raw_df["Date"] >= p_start) &
        (raw_df["Date"] <= p_end)
    ]

    n_input_total = raw_df_period["site"].nunique()

    # Only override if metadata does NOT already define it
    if not str(metadata_get(metadata, "Number of wells", "")).strip():
        metadata["Number of wells"] = str(n_input_total)
    


    site_coords = (
        pd.read_csv(coord_file)
        .drop_duplicates(subset="site")
        .rename(columns={"site": "Well"})
    )

    value_col, inferred_type = detect_value_column_and_type(raw_df)

    if forced_measurement_type in ("depth", "elevation"):
        value_type = forced_measurement_type
        print(f"Detected value column: {value_col} | Measurement type FORCED by metadata: {value_type}")
    else:
        value_type = inferred_type
        print(f"Detected value column: {value_col} | Measurement type inferred: {value_type}")

    # --------------------------------------------------------
    # D) PIVOT TO WIDE + MONTHLY MEANS (daily/monthly input ok)
    # --------------------------------------------------------
    ts_long = raw_df[["Date", "site", value_col]].copy()
    ts_long.rename(columns={"site": "Well", value_col: "Value"}, inplace=True)

    ts_wide = ts_long.pivot_table(index="Date", columns="Well", values="Value", aggfunc="mean").sort_index()
    ts_monthly = ts_wide.resample("ME").mean()
    ts_monthly.index.name = "Date"

    # --------------------------------------------------------
    # E) COMPLETENESS FILTER
    # --------------------------------------------------------
    accepted_df, _rejected_df = New_method_filter(
        ts_monthly,
        percent=completeness_percent,
        start=period_start,
        end=period_end,
        output_excel_path=str(rejected_excel_path),
    )

    if accepted_df.empty or accepted_df.shape[1] == 0:
        raise RuntimeError("No wells accepted after completeness filtering. Check period/thresholds/data.")

    accepted_df.to_csv(accepted_csv_path, index=True, encoding="utf-8")
    print(f"Accepted wells monthly data saved to: {accepted_csv_path}")

    # --------------------------------------------------------
    # F) MK 20y + MK slopes 10/5y
    # --------------------------------------------------------
    mk20 = MK_tests(accepted_df, value_type=value_type, alpha=alpha)
    mk_105 = compute_mk_slopes_fixed_windows_monthly(
        accepted_df, value_type=value_type, alpha=alpha, percent=completeness_percent
    )
    df_mk = mk20.merge(mk_105, on="Well", how="left")

    # --------------------------------------------------------
    # G) ADD COORDINATES + COMPUTE CRI
    # --------------------------------------------------------
    df_mk = df_mk.merge(site_coords[["Well", "X", "Y"]], on="Well", how="left")

    data_for_cri = -accepted_df.copy() if value_type == "depth" else accepted_df.copy()
    iqr = data_for_cri.quantile(0.75) - data_for_cri.quantile(0.25)
    iqr = iqr.replace(0, np.nan)

    df_mk["IQR"] = df_mk["Well"].map(iqr)
    df_mk["CRI"] = df_mk["slope_total_period"] / df_mk["IQR"]

    # --------------------------------------------------------
    # H) EXPORT MK RESULTS
    # --------------------------------------------------------
    df_mk.to_excel(mk_results_path, index=False)
    print(f"MK results saved to: {mk_results_path}")

    # --------------------------------------------------------
    # I) HYDROGRAPH PNG (for Excel sheet)
    # --------------------------------------------------------
    data_for_hydro_excel = -accepted_df.copy() if value_type == "depth" else accepted_df.copy()
    data_for_hydro_excel = data_for_hydro_excel.dropna(how="all", axis=1)

    norm_df_excel = data_for_hydro_excel.sub(data_for_hydro_excel.mean(axis=0), axis=1)
    mean_norm_excel = norm_df_excel.mean(axis=1).to_frame(name="mean_norm").sort_index()

    mean_slopes_year_excel = {
        20: df_mk["mk_slope_20yr_m_per_year"].mean(skipna=True),
        10: df_mk["mk_slope_10yr_m_per_year"].mean(skipna=True),
        5: df_mk["mk_slope_5yr_m_per_year"].mean(skipna=True),
    }
    counts_year_excel = {
        20: df_mk["mk_slope_20yr_m_per_year"].notna().sum(),
        10: df_mk["mk_slope_10yr_m_per_year"].notna().sum(),
        5: df_mk["mk_slope_5yr_m_per_year"].notna().sum(),
    }
    periods_excel = {
        20: (pd.Timestamp("2004-01-01"), pd.Timestamp("2023-12-31")),
        10: (pd.Timestamp("2014-01-01"), pd.Timestamp("2023-12-31")),
        5: (pd.Timestamp("2019-01-01"), pd.Timestamp("2023-12-31")),
    }

    fig_png, ax = plt.subplots(figsize=(12, 6))
    ax.plot(norm_df_excel.index, norm_df_excel, color="#B0B0B0", alpha=0.15, linewidth=0.8)
    ax.plot(mean_norm_excel.index, mean_norm_excel["mean_norm"], color="black", linewidth=2.5,
            label="Mean groundwater level (all wells, monthly)")

    colors_excel = {20: "b", 10: "g", 5: "r"}
    for yrs, (p_start, p_end) in periods_excel.items():
        slope_yr = mean_slopes_year_excel.get(yrs, np.nan)
        if not np.isfinite(slope_yr):
            continue
        sub = mean_norm_excel[(mean_norm_excel.index >= p_start) & (mean_norm_excel.index <= p_end)]
        if len(sub) < 2:
            continue
        y_trend = build_timeaware_trendline(sub.index, sub["mean_norm"].to_numpy(), slope_yr)
        if y_trend is None:
            continue
        n_wells_period = counts_year_excel.get(yrs, 0)
        ax.plot(sub.index, y_trend, linestyle="--", color=colors_excel[yrs], linewidth=2,
                label=f"{yrs}-yr MK Sen slope ({slope_yr:.3f} {forced_units}/yr, n={n_wells_period})")

    ax.set_xlabel("Year")
    ax.set_ylabel("Normalized water table (m)")
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig_png.savefig(hydrograph_png, dpi=300, bbox_inches="tight")
    plt.close(fig_png)

    try:
        wb = load_workbook(mk_results_path)
        ws_h = wb.create_sheet("Hydrograph")
        img = XLImage(str(hydrograph_png))
        ws_h.add_image(img, "A1")
        wb.save(mk_results_path)
        print("Hydrograph image added to sheet 'Hydrograph' in MK_results.xlsx")
    except Exception as e:
        print(f"Could not embed hydrograph image into Excel: {e}")

    # --------------------------------------------------------
    # J) BUILD DASHBOARD HTML (2 pages)
    # --------------------------------------------------------
    units_str = forced_units if forced_units else "m"

    build_dashboard(
        df_mk=df_mk,
        accepted_df=accepted_df,
        value_type=value_type,
        units_str=units_str,
        output_html=output_html,
        region_name=base_path.name,
        metadata=metadata,
        period_start=period_start,
        period_end=period_end,
        completeness_percent=completeness_percent,
        alpha=alpha,
    )

    print("Done.")


# In[ ]:




